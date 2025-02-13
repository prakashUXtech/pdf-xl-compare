"""
File processing service for handling file operations.
"""
from pathlib import Path
import pandas as pd
import streamlit as st
import hashlib
from typing import Dict, Optional, Tuple
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime

class FileProcessor:
    """Handles file operations for the application."""
    
    @staticmethod
    @st.cache_data
    def get_pdf_hash(pdf_content: bytes) -> str:
        """Generate a hash for PDF content."""
        return hashlib.md5(pdf_content).hexdigest()
    
    @staticmethod
    def save_temp_file(file_content: bytes, file_path: Path) -> None:
        """Save content to a temporary file."""
        with open(file_path, "wb") as f:
            f.write(file_content)
    
    @staticmethod
    def load_excel_data(file_path: Path) -> pd.DataFrame:
        """Load and preprocess Excel data."""
        try:
            # Read the Excel file
            excel_file = pd.ExcelFile(file_path)
            
            # Look for the "Wage Entry" sheet first
            target_sheet = "Wage Entry"
            if target_sheet not in excel_file.sheet_names:
                # If not found, try to find a sheet with wage data
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)  # Skip UAN and Name rows
                    if 'WAGES' in df.columns and ('DUE MONTH' in df.columns or 'DUE_MONTH' in df.columns):
                        target_sheet = sheet_name
                        break
            
            # Read the target sheet
            df = pd.read_excel(file_path, sheet_name=target_sheet, skiprows=2)
            
            if df.empty:
                st.error("No data found in the Excel sheet")
                return pd.DataFrame()
            
            # Clean up column names
            df.columns = [str(col).strip().upper() for col in df.columns]
            
            # Select required columns - now using DUE MONTH instead of MONTH
            if 'DUE MONTH' in df.columns and 'WAGES' in df.columns:
                result_df = df[['DUE MONTH', 'WAGES']].copy()
            elif 'DUE_MONTH' in df.columns and 'WAGES' in df.columns:
                result_df = df[['DUE_MONTH', 'WAGES']].copy()
            else:
                st.error("Could not find required columns (DUE MONTH and WAGES) in the Excel sheet")
                return pd.DataFrame()
            
            # Clean up the data
            # Convert month strings to standard format
            result_df['MONTH'] = pd.to_datetime(result_df['DUE MONTH' if 'DUE MONTH' in result_df.columns else 'DUE_MONTH'], format='%b-%y')
            
            # Convert dates to standard string format
            result_df['CREDIT_MONTH'] = result_df['MONTH'].dt.strftime('%m/%Y')
            result_df = result_df.drop(columns=['MONTH'])
            
            # Drop the original DUE MONTH column
            result_df = result_df.drop(columns=['DUE MONTH' if 'DUE MONTH' in result_df.columns else 'DUE_MONTH'])
            
            # Convert wages to numeric, removing any non-numeric characters
            result_df['WAGES'] = pd.to_numeric(
                result_df['WAGES'].astype(str).str.replace(r'[^\d.-]', '', regex=True), 
                errors='coerce'
            )
            
            # Drop rows where either column is null
            result_df = result_df.dropna()
            
            if not result_df.empty:
                st.success(f"Successfully loaded data from sheet '{target_sheet}'")
                return result_df
            
            st.error("Could not find required columns (DUE MONTH and WAGES) in the Excel sheet")
            return pd.DataFrame()
            
        except Exception as e:
            st.error(f"Error loading Excel file: {str(e)}")
            return pd.DataFrame()
    
    @staticmethod
    @st.cache_data(show_spinner=False)
    def load_processed_files() -> Dict[str, pd.DataFrame]:
        """Load all available processed CSV files."""
        processed_files = {}
        processed_dir = Path("data/processed")
        
        if not processed_dir.exists():
            st.warning(f"Directory not found: {processed_dir}")
            return processed_files
        
        try:
            csv_files = list(processed_dir.glob("*.csv"))
            if not csv_files:
                st.warning(f"No CSV files found in {processed_dir}")
                return processed_files
            
            for file in csv_files:
                try:
                    df = pd.read_csv(file)
                    if df.empty:
                        continue
                    
                    # Clean and standardize
                    df = FileProcessor._standardize_dataframe(df)
                    
                    if FileProcessor._validate_dataframe(df):
                        processed_files[file.name] = df
                        st.success(f"Successfully loaded {file.name}")
                
                except Exception as e:
                    st.error(f"Error loading {file.name}: {str(e)}")
                    continue
        
        except Exception as e:
            st.error(f"Error accessing directory: {str(e)}")
            return processed_files
        
        if processed_files:
            st.success(f"Successfully loaded {len(processed_files)} files")
        else:
            st.warning("No valid files found with required columns")
        
        return processed_files
    
    @staticmethod
    def _standardize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Standardize DataFrame columns and format."""
        # Clean column names
        df.columns = [str(col).strip().upper() if isinstance(col, (str, int, float)) else str(col).upper() 
                     for col in df.columns]
        
        # Handle different file formats
        if 'WAGES_CONTRIB' in df.columns:
            df = df.rename(columns={'WAGES_CONTRIB': 'WAGES'})
        elif 'WAGES_EXCEL' in df.columns:
            df = df.rename(columns={'WAGES_EXCEL': 'WAGES'})
        
        # Clean credit month format
        if 'CREDIT_MONTH' in df.columns:
            df['CREDIT_MONTH'] = df['CREDIT_MONTH'].astype(str).str.strip()
        
        # Convert wages to numeric
        wages_cols = [col for col in ['WAGES', 'WAGES_CONTRIB', 'WAGES_EXCEL'] if col in df.columns]
        if wages_cols:
            df[wages_cols[0]] = pd.to_numeric(df[wages_cols[0]], errors='coerce')
        
        return df
    
    @staticmethod
    def _validate_dataframe(df: pd.DataFrame) -> bool:
        """Validate if DataFrame has required columns."""
        required_cols = ['CREDIT_MONTH']
        wage_cols = ['WAGES', 'WAGES_CONTRIB', 'WAGES_EXCEL']
        
        has_required = all(col in df.columns for col in required_cols)
        has_wages = any(col in df.columns for col in wage_cols)
        
        return has_required and has_wages
    
    @staticmethod
    def cleanup_temp_files(*file_paths: Path) -> None:
        """Clean up temporary files."""
        for path in file_paths:
            if path.exists():
                try:
                    path.unlink()
                except Exception as e:
                    st.error(f"Error cleaning up {path}: {str(e)}")
    
    @staticmethod
    def export_comparison_with_highlights(comparison_df: pd.DataFrame) -> bytes:
        """
        Export comparison results to Excel with highlighted mismatches.
        
        Args:
            comparison_df: DataFrame containing comparison results with MISMATCH column
            
        Returns:
            bytes: Excel file content as bytes
        """
        # Create a copy to avoid modifying the original
        df = comparison_df.copy()
        
        # Create a Pandas Excel writer using openpyxl
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Convert the dataframe to an XlsxWriter Excel object
            df.to_excel(writer, sheet_name='Comparison Results', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Comparison Results']
            
            # Define the format for mismatched cells
            fill = PatternFill(start_color='FFE6E6',
                             end_color='FFE6E6',
                             fill_type='solid')
            
            # Get the dimensions of the dataframe
            rows = len(df) + 1  # +1 for header
            cols = len(df.columns)
            
            # Apply conditional formatting based on MISMATCH column
            for row in range(2, rows + 1):  # Start from 2 to skip header
                if df.iloc[row-2]['MISMATCH']:
                    # Highlight WAGES_PDF and WAGES_EXCEL cells for mismatched rows
                    pdf_col = df.columns.get_loc('WAGES_PDF') + 1
                    excel_col = df.columns.get_loc('WAGES_EXCEL') + 1
                    
                    worksheet.cell(row=row, column=pdf_col).fill = fill
                    worksheet.cell(row=row, column=excel_col).fill = fill
            
            # Auto-adjust columns width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Reset pointer and return bytes
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def highlight_mismatches_in_original_excel(
        excel_file_content: bytes,
        comparison_df: pd.DataFrame,
        month_col: str,
        wages_col: str
    ) -> bytes:
        """
        Highlight mismatches in the original Excel file while preserving its structure.
        
        Args:
            excel_file_content: Original Excel file content as bytes
            comparison_df: DataFrame with comparison results
            month_col: Name of the month column in original Excel
            wages_col: Name of the wages column in original Excel
            
        Returns:
            bytes: Modified Excel file content with highlighted cells
        """
        # Create a temporary file to work with
        output = BytesIO()
        output.write(excel_file_content)
        output.seek(0)
        
        # Load the workbook
        wb = load_workbook(output)
        
        # Try to find the Wage Entry sheet
        if "Wage Entry" in wb.sheetnames:
            ws = wb["Wage Entry"]
        else:
            ws = wb.active
        
        # Get mismatched months and their corresponding wages
        mismatches = comparison_df[comparison_df['MISMATCH']]['CREDIT_MONTH'].tolist()
        
        # Find column indices - start from row 3 (1-based index) to skip headers
        header_row = 3  # Skip UAN and Name rows
        month_idx = None
        wages_idx = None
        
        # Find the month and wages columns
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(header_row, col).value or '').strip().upper()
            if cell_value == 'MONTH':
                month_idx = col
            elif cell_value == 'WAGES':
                wages_idx = col
            
            if month_idx and wages_idx:
                break
        
        if not (month_idx and wages_idx):
            raise ValueError("Could not find MONTH or WAGES columns in Excel file")
        
        # Define the highlight style
        highlight_fill = PatternFill(
            start_color='FFE6E6',
            end_color='FFE6E6',
            fill_type='solid'
        )
        
        # Iterate through rows and highlight mismatches (start from row after header)
        for row in range(header_row + 1, ws.max_row + 1):
            month_cell = ws.cell(row, month_idx)
            wages_cell = ws.cell(row, wages_idx)
            
            # Convert cell date to string format matching comparison_df
            try:
                if isinstance(month_cell.value, datetime):
                    month_str = month_cell.value.strftime('%m/%Y')
                else:
                    # Try to parse the date string
                    month_str = pd.to_datetime(str(month_cell.value)).strftime('%m/%Y')
                
                if month_str in mismatches:
                    wages_cell.fill = highlight_fill
            except:
                continue  # Skip if date parsing fails
        
        # Save the modified workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def fix_mismatches_in_original_excel(self, original_excel_bytes: bytes, comparison: pd.DataFrame, month_col: str, wages_col: str) -> bytes:
        """Fix mismatched wage cells in the original Excel file using PDF values from the comparison DataFrame while preserving all sheets and formatting."""
        import pandas as pd
        from io import BytesIO
        from openpyxl import load_workbook

        # Build mapping of month (formatted as 'MM/YYYY') to correct wage from PDF
        fix_mapping = {}
        for _, row in comparison.iterrows():
            if row.get('MISMATCH') and pd.notnull(row.get('WAGES_PDF')):
                fix_mapping[row['CREDIT_MONTH']] = row['WAGES_PDF']

        # Load the workbook from bytes
        wb = load_workbook(filename=BytesIO(original_excel_bytes))
        # Assume the relevant data is in the active sheet
        ws = wb.active

        # Get header row (assumed to be the first row) to locate columns
        header = [cell.value for cell in ws[1]]
        month_index = None
        wages_index = None
        for idx, col in enumerate(header, start=1):
            if col and month_col.upper() in str(col).upper():
                month_index = idx
            if col and wages_col.upper() in str(col).upper():
                wages_index = idx

        if month_index is None or wages_index is None:
            raise Exception("Unable to find required columns in the Excel file.")

        # Iterate over the data rows (assuming header is in row 1)
        for row in ws.iter_rows(min_row=2):
            cell_month = row[month_index - 1]
            cell_wages = row[wages_index - 1]
            try:
                # Attempt to parse the month cell value into a date with format 'MM/YYYY'
                date_val = pd.to_datetime(str(cell_month.value).strip(), format='%m/%Y', errors='coerce')
                if pd.notnull(date_val):
                    month_str = date_val.strftime('%m/%Y')
                    if month_str in fix_mapping:
                        cell_wages.value = fix_mapping[month_str]
            except Exception:
                pass

        # Save workbook back to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    @staticmethod
    def fix_highlighted_excel(highlighted_excel_bytes: bytes, comparison: pd.DataFrame) -> bytes:
        """Fix highlighted cells in Excel file using PDF values from comparison DataFrame."""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from io import BytesIO
        import pandas as pd
        
        # Load the workbook from bytes
        wb = load_workbook(filename=BytesIO(highlighted_excel_bytes))
        ws = wb.active  # Get active sheet
        
        # Create mapping of month to correct wage from PDF
        fix_mapping = {}
        for _, row in comparison.iterrows():
            if row.get('MISMATCH') and pd.notnull(row.get('WAGES_PDF')):
                try:
                    date = pd.to_datetime(row['CREDIT_MONTH'], format='%m/%Y')
                    mmm_yy = date.strftime('%b-%y')  # Format as 'Apr-09'
                    fix_mapping[mmm_yy.upper()] = float(row['WAGES_PDF'])  # Convert to float for consistency
                    print(f"Added to mapping: {mmm_yy.upper()} -> {float(row['WAGES_PDF'])}")
                except Exception as e:
                    print(f"Error converting date {row['CREDIT_MONTH']}: {str(e)}")
                    continue
        
        print("Fix mapping created:", fix_mapping)  # Debug print
        
        # First find the MONTH and WAGES columns
        month_col = None
        wages_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=3, column=col).value or '').upper()
            if 'MON' in cell_value:
                month_col = col
            if 'WAGE' in cell_value and 'S' in cell_value:  # Match 'WAGES' column
                wages_col = col
        
        if not month_col or not wages_col:
            raise Exception(f"Could not find required columns. Month col: {month_col}, Wages col: {wages_col}")
        
        print(f"Found columns - Month: {month_col}, Wages: {wages_col}")  # Debug print
        
        # Iterate through rows starting from data rows (after headers)
        changes_made = 0
        for row in range(4, ws.max_row + 1):  # Start from row 4 (after headers)
            month_cell = ws.cell(row=row, column=month_col)
            wage_cell = ws.cell(row=row, column=wages_col)
            
            # Debug print cell fill information
            print(f"Row {row} - Fill type: {wage_cell.fill.fill_type}, Pattern: {wage_cell.fill.patternType}, Color: {wage_cell.fill.start_color.rgb}")
            
            # Check if this is a highlighted wage cell - check both fill_type and color
            if (wage_cell.fill.fill_type == 'solid' or 
                wage_cell.fill.patternType == 'solid') and wage_cell.fill.start_color.rgb in ['FFE6E6', 'FFE6E6E6']:
                
                month_str = str(month_cell.value).upper() if month_cell.value else ''
                print(f"Found highlighted cell at row {row} - Month: {month_str}, Current wage: {wage_cell.value}")
                
                if month_str in fix_mapping:
                    old_value = wage_cell.value
                    new_value = fix_mapping[month_str]
                    wage_cell.value = new_value
                    wage_cell.fill = PatternFill(fill_type=None)  # Remove highlight
                    changes_made += 1
                    print(f"FIXED: Row {row} - Month {month_str}: {old_value} -> {new_value}")
        
        print(f"Total changes made: {changes_made}")
        
        if changes_made == 0:
            print("WARNING: No changes were made to the Excel file!")
            print("Please check if the cells are actually highlighted with the expected color (FFE6E6)")
        
        # Save the modified workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def fix_and_highlight_excel(
        excel_file_content: bytes,
        comparison_df: pd.DataFrame,
        month_col: str,
        wages_col: str
    ) -> bytes:
        """
        Fix mismatched cells with PDF values and highlight them in the original Excel file.
        
        Args:
            excel_file_content: Original Excel file content as bytes
            comparison_df: DataFrame with comparison results
            month_col: Name of the month column in original Excel
            wages_col: Name of the wages column in original Excel
            
        Returns:
            bytes: Modified Excel file content with fixed and highlighted cells
        """
        # Create a temporary file to work with
        output = BytesIO()
        output.write(excel_file_content)
        output.seek(0)
        
        # Load the workbook
        wb = load_workbook(output)
        
        # Try to find the Wage Entry sheet
        if "Wage Entry" in wb.sheetnames:
            ws = wb["Wage Entry"]
        else:
            ws = wb.active
        
        # Get mismatched months and create a mapping of corrections
        mismatches = comparison_df[comparison_df['MISMATCH']]
        corrections = {
            row['CREDIT_MONTH']: row['WAGES_PDF']
            for _, row in mismatches.iterrows()
            if pd.notnull(row['WAGES_PDF'])
        }
        
        # Find column indices - start from row 3 (1-based index) to skip headers
        header_row = 3  # Skip UAN and Name rows
        month_idx = None
        wages_idx = None
        
        # Find the month and wages columns
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(header_row, col).value or '').strip().upper()
            if cell_value == 'MONTH':
                month_idx = col
            elif cell_value == 'WAGES':
                wages_idx = col
            
            if month_idx and wages_idx:
                break
        
        if not (month_idx and wages_idx):
            raise ValueError("Could not find MONTH or WAGES columns in Excel file")
        
        # Define the highlight style
        highlight_fill = PatternFill(
            start_color='FFE6E6',
            end_color='FFE6E6',
            fill_type='solid'
        )
        
        # Iterate through rows and highlight mismatches (start from row after header)
        for row in range(header_row + 1, ws.max_row + 1):
            month_cell = ws.cell(row, month_idx)
            wages_cell = ws.cell(row, wages_idx)
            
            # Convert cell date to string format matching comparison_df
            try:
                if isinstance(month_cell.value, datetime):
                    month_str = month_cell.value.strftime('%m/%Y')
                else:
                    # Try to parse the date string
                    month_str = pd.to_datetime(str(month_cell.value)).strftime('%m/%Y')
                
                if month_str in corrections:
                    # Apply correction
                    wages_cell.value = corrections[month_str]
                    # Apply highlighting
                    wages_cell.fill = highlight_fill
            except:
                continue  # Skip if date parsing fails
        
        # Save the modified workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue() 