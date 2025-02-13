"""
Integration tests for raw response processing.
"""
import json
from pathlib import Path
import pandas as pd
from app.services.ocr.nanonets import NanonetsOCR
from io import BytesIO
from openpyxl import load_workbook

def test_raw_response():
    """Test data extraction from existing raw response."""
    print("Testing data extraction from raw response...")
    
    # Find the latest raw response file
    raw_responses_dir = Path('raw_responses')
    if not raw_responses_dir.exists():
        print("No raw_responses directory found!")
        return
    
    response_files = list(raw_responses_dir.glob('raw_response_*.json'))
    if not response_files:
        print("No raw response files found!")
        return
    
    # Get the latest file
    latest_file = max(response_files, key=lambda x: x.stat().st_mtime)
    print(f"\nFound raw response file: {latest_file}")
    
    try:
        # Load the raw response
        with open(latest_file) as f:
            raw_response = json.load(f)
        
        print("\nSuccessfully loaded raw response")
        
        # Initialize OCR service
        ocr = NanonetsOCR("test_key", "test_model")
        
        # Extract table data
        print("\nExtracting table data...")
        df = ocr._process_raw_response(raw_response)
        
        if df is not None and not df.empty:
            print("\nSuccessfully extracted data!")
            print("\nData Preview:")
            print(df.head())
            print("\nColumns:", df.columns.tolist())
            print(f"\nTotal rows: {len(df)}")
            
            # Basic data analysis
            if 'WAGES' in df.columns:
                print("\nWages Statistics:")
                print(df['WAGES'].describe())
            
            if 'CREDIT_MONTH' in df.columns:
                print("\nDate Range:")
                print(f"From: {df['CREDIT_MONTH'].min()}")
                print(f"To: {df['CREDIT_MONTH'].max()}")
            
            # Save extracted data
            output_file = Path('data/processed/test_extracted_data.csv')
            output_file.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(output_file, index=False)
            print(f"\nSaved extracted data to: {output_file}")
        else:
            print("No valid data extracted from the response")
        
    except Exception as e:
        print(f"Error processing raw response: {str(e)}")

def test_fix_and_highlight_excel(self):
    """Test fixing and highlighting mismatched cells in Excel."""
    # Create a simple test Excel file
    df = pd.DataFrame({
        'MONTH': ['Jan-23', 'Feb-23', 'Mar-23'],
        'WAGES': [1000, 2000, 3000]
    })
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_content = excel_buffer.getvalue()
    
    # Create a comparison DataFrame with mismatches
    comparison_df = pd.DataFrame({
        'CREDIT_MONTH': ['01/2023', '02/2023', '03/2023'],
        'WAGES_PDF': [1500, 2000, 3500],
        'WAGES_EXCEL': [1000, 2000, 3000],
        'MISMATCH': [True, False, True]
    })
    
    # Process the Excel file
    processor = FileProcessor()
    result = processor.fix_and_highlight_excel(
        excel_content,
        comparison_df,
        'MONTH',
        'WAGES'
    )
    
    # Load the result back into a workbook
    wb = load_workbook(BytesIO(result))
    ws = wb.active
    
    # Check that values were corrected and cells were highlighted
    assert ws['B2'].value == 1500  # First row should be corrected
    assert ws['B2'].fill.start_color.rgb == 'FFE6E6'  # Should be highlighted
    assert ws['B3'].value == 2000  # Second row should be unchanged
    assert ws['B3'].fill.start_color.rgb == '00000000'  # Should not be highlighted
    assert ws['B4'].value == 3500  # Third row should be corrected
    assert ws['B4'].fill.start_color.rgb == 'FFE6E6'  # Should be highlighted

if __name__ == '__main__':
    test_raw_response() 